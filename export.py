import pandas as pd #type: ignore
# from fpdf import FPDF #type: ignore
from reportlab.lib.pagesizes import A4 #type: ignore
from reportlab.lib import colors #type: ignore
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle #type: ignore
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, KeepTogether #type: ignore
from reportlab.pdfgen import canvas #type: ignore
from reportlab.lib.units import cm #type: ignore
from datetime import datetime
from openpyxl import Workbook #type: ignore
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font #type: ignore
from openpyxl.utils.dataframe import dataframe_to_rows #type: ignore

# PDF Class for easier code use 
class PDFReport:
    def __init__(self, filename):
        self.filename = filename
        self.doc = SimpleDocTemplate(filename, pagesize=A4)
        self.elements = []
        
    def add_footer(self, canvas, doc):
        footer_left = "www.agf.co.id  |  office@agf.co.id"
        footer_right = f"Generated on: {datetime.now().strftime('%d %B %Y')}"
        
        canvas.saveState()
        canvas.setFont("Helvetica", 8)
        canvas.drawString(10, 20, footer_left)  # Left-aligned footer text
        canvas.drawRightString(A4[0] - 10, 20, footer_right)  # Right-aligned footer text
        canvas.restoreState()
        
    def build_pdf(self):
        # Build the document with footer on each page
        self.doc.build(self.elements, onFirstPage=self.add_footer, onLaterPages=self.add_footer)

def generate_balance_sheet_pdf(balance_sheet_data, filename, language, report_date):
    pdf = PDFReport(filename)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(name='Title', fontSize=16, alignment=1, spaceAfter=20, fontName='Helvetica-Bold')
    
    title = "Balance Sheet" if language != 'i' else "Neraca"
    pdf.elements.append(Paragraph(title, title_style))
    
    # Add report date
    date_text = f"As of: {report_date.strftime('%d %B %Y')}" if language != 'i' else f"Per tanggal: {report_date.strftime('%d %B %Y')}"
    pdf.elements.append(Paragraph(date_text, styles['Normal']))
    pdf.elements.append(Spacer(1, 12))
    
    # # Assets
    # data.append(["Assets" if language != 'i' else "Aset", ""])
    # for asset, value in balance_sheet_data['assets'].items():
    #     data.append([f"  {asset}", f"Rp {value:,.2f}"])
    
    # # Liabilities
    # data.append(["Liabilities" if language != 'i' else "Kewajiban", ""])
    # for liability, value in balance_sheet_data['liabilities'].items():
    #     data.append([f"  {liability}", f"Rp {value:,.2f}"])
    
    # # Equity
    # data.append(["Equity" if language != 'i' else "Ekuitas", ""])
    # for equity_item, value in balance_sheet_data['equity'].items():
    #     data.append([f"  {equity_item}", f"Rp {value:,.2f}"])
    
    # # Total Liabilities and Equity
    # data.append(["Total Liabilities and Equity" if language != 'i' else "Total Kewajiban dan Ekuitas", 
    #              f"Rp {balance_sheet_data['total_liabilities_and_equity']:,.2f}"])
    
    # for section, items in balance_sheet_data.items():
    #     if section != 'total_liabilities_and_equity':
    #         section_title = section.capitalize() if language != 'i' else section.capitalize().translate(str.maketrans("ABCDEFGHIJKLMNOPQRSTUVWXYZ", "ABCDEFGHIJKLMNOPQRSTUVWXYZ".lower()))
    #         data.append([section_title, ""])
    #         for item, value in items.items():
    #             if isinstance(value, dict):
    #                 for sub_item, sub_value in value.items():
    #                     data.append([f"  {sub_item}", f"Rp {sub_value:,.2f}"])
    #             else:
    #                 data.append([f"  {item}", f"Rp {value:,.2f}"])

    # data.append(["Total Liabilities and Equity" if language != 'i' else "Total Kewajiban dan Ekuitas", 
    #              f"Rp {balance_sheet_data['total_liabilities_and_equity']:,.2f}"])

    data = []
    
    # Assets
    data.append(["Assets" if language != 'i' else "Aset", ""])
    for asset_type, asset_values in balance_sheet_data['assets'].items():
        if isinstance(asset_values, dict):
            data.append([f"  {asset_type}", ""])
            for sub_item, sub_value in asset_values.items():
                data.append([f"    {sub_item}", f"Rp {sub_value:,.2f}"])
        else:
            data.append([f"  {asset_type}", f"Rp {asset_values:,.2f}"])
    
    # Liabilities and Equity
    data.append(["Liabilities and Equity" if language != 'i' else "Kewajiban dan Ekuitas", ""])
    for section, items in balance_sheet_data['liabilities_and_equity'].items():
        if isinstance(items, dict):
            data.append([f"  {section}", ""])
            for sub_item, sub_value in items.items():
                data.append([f"    {sub_item}", f"Rp {sub_value:,.2f}"])
        else:
            data.append([f"  {section}", f"Rp {items:,.2f}"])
    
    table = Table(data, colWidths=[300, 200])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black)
    ]))
    
    pdf.elements.append(table)
    pdf.build_pdf()

def generate_pnl_pdf(revenues, expenses, net_income, filename, language, report_date):
    pdf = PDFReport(filename)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(name='Title', fontSize=16, alignment=1, spaceAfter=20, fontName='Helvetica-Bold')
    
    title = "Profit and Loss Statement" if language != 'i' else "Laporan Laba Rugi"
    pdf.elements.append(Paragraph(title, title_style))
    
    # Add report date
    date_text = f"Date: {report_date.strftime('%d %B %Y')}" if language != 'i' else f"Tanggal: {report_date.strftime('%d %B %Y')}"
    pdf.elements.append(Paragraph(date_text, styles['Normal']))
    pdf.elements.append(Spacer(1, 12))
    
    data = [["Item", "Amount"]]
    
    # Add revenues
    for revenue, value in revenues.items():
        data.append([f"{revenue}", f"Rp {value:,.2f}"])
    
    data.append(["Total Revenue" if language != 'i' else "Total Pendapatan", f"Rp {sum(revenues.values()):,.2f}"])
    
    # Add expenses
    for expense, value in expenses.items():
        data.append([f"{expense}", f"Rp {value:,.2f}"])
    
    data.append(["Total Expenses" if language != 'i' else "Total Pengeluaran", f"Rp {sum(expenses.values()):,.2f}"])
    data.append(["Net Income" if language != 'i' else "Laba Bersih", f"Rp {net_income:,.2f}"])
    
    table = Table(data, colWidths=[300, 200])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black)
    ]))
    
    pdf.elements.append(table)
    pdf.build_pdf()

def generate_cash_flow_pdf(operating_activities, investing_activities, financing_activities, net_cash_flow, beginning_cash, ending_cash, filename, language, report_date):
    pdf = PDFReport(filename)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(name='Title', fontSize=16, alignment=1, spaceAfter=20, fontName='Helvetica-Bold')
    
    title = "Cash Flow Statement" if language != 'i' else "Laporan Arus Kas"
    pdf.elements.append(Paragraph(title, title_style))
    
    # Add report date
    date_text = f"Date: {report_date.strftime('%d %B %Y')}" if language != 'i' else f"Tanggal: {report_date.strftime('%d %B %Y')}"
    pdf.elements.append(Paragraph(date_text, styles['Normal']))
    pdf.elements.append(Spacer(1, 12))
    
    data = [["Activity", "Amount"]]
    
    # Operating Activities
    data.append(["Operating Activities" if language != 'i' else "Aktivitas Operasi", ""])
    for activity, value in operating_activities.items():
        data.append([f"  {activity}", f"Rp {value:,.2f}"])
    data.append(["Net Cash from Operating Activities" if language != 'i' else "Arus Kas Bersih dari Aktivitas Operasi", 
                 f"Rp {sum(operating_activities.values()):,.2f}"])
    
    # Investing Activities
    data.append(["Investing Activities" if language != 'i' else "Aktivitas Investasi", ""])
    for activity, value in investing_activities.items():
        data.append([f"  {activity}", f"Rp {value:,.2f}"])
    data.append(["Net Cash from Investing Activities" if language != 'i' else "Arus Kas Bersih dari Aktivitas Investasi", 
                 f"Rp {sum(investing_activities.values()):,.2f}"])
    
    # Financing Activities
    data.append(["Financing Activities" if language != 'i' else "Aktivitas Pendanaan", ""])
    for activity, value in financing_activities.items():
        data.append([f"  {activity}", f"Rp {value:,.2f}"])
    data.append(["Net Cash from Financing Activities" if language != 'i' else "Arus Kas Bersih dari Aktivitas Pendanaan", 
                 f"Rp {sum(financing_activities.values()):,.2f}"])
    
    # Net Cash Flow
    data.append(["Net Cash Flow" if language != 'i' else "Arus Kas Bersih", f"Rp {net_cash_flow:,.2f}"])
    
    # Beginning and Ending Cash
    data.append(["Beginning Cash" if language != 'i' else "Kas Awal", f"Rp {beginning_cash:,.2f}"])
    data.append(["Ending Cash" if language != 'i' else "Kas Akhir", f"Rp {ending_cash:,.2f}"])
    
    table = Table(data, colWidths=[300, 200])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black)
    ]))
    
    pdf.elements.append(table)
    pdf.build_pdf()

def generate_equity_statement_pdf(equity_components, total_equity, filename, language, report_date):
    pdf = PDFReport(filename)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(name='Title', fontSize=16, alignment=1, spaceAfter=20, fontName='Helvetica-Bold')
    
    title = "Statement of Changes in Equity" if language != 'i' else "Laporan Perubahan Ekuitas"
    pdf.elements.append(Paragraph(title, title_style))
    
    # Add report date
    date_text = f"Date: {report_date.strftime('%d %B %Y')}" if language != 'i' else f"Tanggal: {report_date.strftime('%d %B %Y')}"
    pdf.elements.append(Paragraph(date_text, styles['Normal']))
    pdf.elements.append(Spacer(1, 12))
    
    data = [["Component" if language != 'i' else "Komponen", "Amount" if language != 'i' else "Jumlah"]]
    
    for component, value in equity_components.items():
        data.append([component, f"Rp {value:,.2f}"])
    
    data.append(["Total Equity" if language != 'i' else "Total Ekuitas", f"Rp {total_equity:,.2f}"])
    
    table = Table(data, colWidths=[300, 200])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black)
    ]))
    
    pdf.elements.append(table)
    pdf.build_pdf()

def generate_pdf(summary, total_sales, total_purchase, net_profit, final_capital, total_assets, total_liabilities, total_equity, filename, language, report_date, report_type):
    pdf = PDFReport(filename)  # Create an instance of PDFReport
    
    # Custom styles
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='CenterTitle', alignment=1, fontSize=16, spaceAfter=20, fontName='Helvetica-Bold'))

    # Format header date based on report type
    if report_type == "yearly":
        header_text = f"Date of report for the year {report_date.year}" if language != 'i' else f"Tanggal Laporan untuk tahun {report_date.year}"
    elif report_type == "monthly":
        month_name = report_date.strftime('%B')
        header_text = f"Date of report for {month_name} {report_date.year}" if language != 'i' else f"Tanggal Laporan untuk {month_name} {report_date.year}"
    elif report_type == "weekly":
        start_of_week = report_date.strftime('%d %B %Y')
        end_of_week = (report_date + pd.Timedelta(days=6)).strftime('%d %B %Y')
        header_text = f"Date of report for the week of {start_of_week} to {end_of_week}" if language != 'i' else f"Tanggal Laporan untuk minggu {start_of_week} sampai {end_of_week}"
    else:
        header_text = "Date of report: " + report_date.strftime('%d %B %Y')

    pdf.elements.append(Paragraph(header_text, styles['CenterTitle']))

    # Column names and titles based on language
    column_names = ["Komponen", "Product ID", "Jumlah"] if language == 'i' else ["Component", "Product ID", "Amount"]
    data = [column_names]
    
    # Add rows from summary
    for product_id, amount in summary.get('sales', {}).items():
        data.append(["Jasa Pendapatan", product_id, f"Rp {amount:,.2f}"])

    # Add Total row
    data.append(["Total Pendapatan", "", f"Rp {total_sales:,.2f}"])

    # Create the table and style it
    table = Table(data, colWidths=[150, 100, 120])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black)
    ]))
    pdf.elements.append(table)
    pdf.elements.append(Spacer(1, 20))

    # Financial Summary section
    financial_data = [
        ("Laba Rugi Harian" if language == 'i' else "Daily Profit and Loss", f"Rp {net_profit:,.2f}"),
        ("Total Arus Kas" if language == 'i' else "Total Cash Flow", f"Rp {net_profit:,.2f}"),
        ("Saldo Kas Harian" if language == 'i' else "Daily Cash Balance", f"Rp {final_capital:,.2f}"),
        ("Total Aset" if language == 'i' else "Total Assets", f"Rp {total_assets:,.2f}"),
        ("Total Utang" if language == 'i' else "Total Debt", f"Rp {total_liabilities:,.2f}"),
        ("Total Ekuitas" if language == 'i' else "Total Equity", f"Rp {total_equity:,.2f}")
    ]

    # Add Financial Summary as a separate table
    summary_table_data = [[title, value] for title, value in financial_data]
    summary_table = Table(summary_table_data, colWidths=[250, 120])
    summary_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('BACKGROUND', (0, 0), (-1, -1), colors.whitesmoke),
    ]))
    pdf.elements.append(summary_table)

    # Build the PDF with the footer
    pdf.build_pdf()

def generate_bilingual_pdf_report(filename, summary, total_sales, total_purchase, net_profit, final_capital, total_assets, total_liabilities, total_equity, report_period):
    pdf = PDFReport(filename)  # Create an instance of PDFReport
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(name='Title', fontSize=16, alignment=1, leading=20)
    header_style = ParagraphStyle(name='Header', fontSize=12, alignment=0, spaceBefore=12, spaceAfter=6)
    
    # Header
    title_id = Paragraph("Laporan Keuangan", title_style)
    title_en = Paragraph("Financial Report", title_style)
    date_id = Paragraph(f"Tanggal Laporan: {report_period}", header_style)
    date_en = Paragraph(f"Date of report: {report_period}", header_style)
    pdf.elements.extend([KeepTogether([title_id, title_en]), Spacer(1, 0.2 * cm), KeepTogether([date_id, date_en]), Spacer(1, 0.5 * cm)])
    
    # Table setup for Indonesian and English columns
    table_data = [["Komponen", "Product ID", "Jumlah", "Component", "Product ID", "Amount"]]
    for product_id, amount in summary.get('sales', {}).items():
        component_id = "Pendapatan Jasa"
        component_en = "Service Revenue"
        table_data.append([
            component_id, product_id, f"Rp {amount:.2f}",
            component_en, product_id, f"Rp {amount:.2f}"
        ])
    
    # Add totals for both Indonesian and English columns
    table_data.append(["Total Pendapatan", "", f"Rp {total_sales:.2f}", "Total Sales", "", f"Rp {total_sales:.2f}"])

    # Create the main table and set styles
    table = Table(table_data, colWidths=[80, 60, 60, 80, 60, 60])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    pdf.elements.append(table)
    pdf.elements.append(Spacer(1, 0.5 * cm))
    
    # Financial Summary in Indonesian and English
    summary_data = [
        ["Laba Rugi Harian", f"Rp {net_profit:.2f}", "Daily Profit and Loss", f"Rp {net_profit:.2f}"],
        ["Total Arus Kas", f"Rp {net_profit:.2f}", "Total Cash Flow", f"Rp {net_profit:.2f}"],
        ["Saldo Kas Harian", f"Rp {final_capital:.2f}", "Daily Cash Balance", f"Rp {final_capital:.2f}"],
        ["Total Aset", f"Rp {total_assets:.2f}", "Total Assets", f"Rp {total_assets:.2f}"],
        ["Total Utang", f"Rp {total_liabilities:.2f}", "Total Debt", f"Rp {total_liabilities:.2f}"],
        ["Total Ekuitas", f"Rp {total_equity:.2f}", "Total Equity", f"Rp {total_equity:.2f}"]
    ]
    
    # Create summary table with styling
    summary_table = Table(summary_data, colWidths=[80, 60, 80, 60])
    summary_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black)
    ]))
    
    pdf.elements.append(summary_table)
    
    # Build PDF with footer
    pdf.build_pdf()

def export_balance_sheet_to_excel(balance_sheet_data, filename, language, report_date):
    wb = Workbook()
    ws = wb.active
    ws.title = "Balance Sheet"
    
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    number_format = '#,##0.00'
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Add title and date
    ws['A1'] = "Balance Sheet" if language != 'i' else "Neraca"
    ws['A2'] = f"As of: {report_date.strftime('%d %B %Y')}" if language != 'i' else f"Per tanggal: {report_date.strftime('%d %B %Y')}"
    ws['A1'].font = header_font
    ws['A2'].font = Font(italic=True)
    
    # Add data
    row = 4
    
    # Assets
    ws.cell(row=row, column=1, value="Assets" if language != 'i' else "Aset")
    ws.cell(row=row, column=1).font = header_font
    row += 1
    
    for asset_type, asset_values in balance_sheet_data['assets'].items():
        if isinstance(asset_values, dict):
            ws.cell(row=row, column=1, value=asset_type)
            ws.cell(row=row, column=1).font = Font(bold=True)
            row += 1
            for sub_item, sub_value in asset_values.items():
                ws.cell(row=row, column=1, value=sub_item)
                ws.cell(row=row, column=2, value=sub_value).number_format = number_format
                row += 1
        else:
            ws.cell(row=row, column=1, value=asset_type)
            ws.cell(row=row, column=2, value=asset_values).number_format = number_format
            row += 1
    
    row += 1  # Add a blank row for separation
    
    # Liabilities and Equity
    ws.cell(row=row, column=1, value="Liabilities and Equity" if language != 'i' else "Kewajiban dan Ekuitas")
    ws.cell(row=row, column=1).font = header_font
    row += 1
    
    for section, items in balance_sheet_data['liabilities_and_equity'].items():
        if isinstance(items, dict):
            ws.cell(row=row, column=1, value=section)
            ws.cell(row=row, column=1).font = Font(bold=True)
            row += 1
            for sub_item, sub_value in items.items():
                ws.cell(row=row, column=1, value=sub_item)
                ws.cell(row=row, column=2, value=sub_value).number_format = number_format
                row += 1
        else:
            ws.cell(row=row, column=1, value=section)
            ws.cell(row=row, column=2, value=items).number_format = number_format
            row += 1
    
    # Apply styles
    for row in ws[f'A4:B{ws.max_row}']:
        for cell in row:
            cell.border = border
    
    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    
    wb.save(filename)


def export_pnl_to_excel(revenues, expenses, net_income, filename, language):
    wb = Workbook()
    ws = wb.active
    ws.title = "Profit and Loss Statement"
    
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    number_format = '#,##0.00'
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Add headers
    headers = ["Item", "Amount"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    # Add revenue data
    for revenue, value in revenues.items():
        ws.append([revenue, value])
        ws.cell(row=ws.max_row, column=2).number_format = number_format
    
    ws.append(["Total Revenue" if language != 'i' else "Total Pendapatan", sum(revenues.values())])
    ws.cell(row=ws.max_row, column=2).number_format = number_format
    ws.cell(row=ws.max_row, column=1).font = header_font
    
    # Add expense data
    for expense, value in expenses.items():
        ws.append([expense, value])
        ws.cell(row=ws.max_row, column=2).number_format = number_format
    
    ws.append(["Total Expenses" if language != 'i' else "Total Pengeluaran", sum(expenses.values())])
    ws.cell(row=ws.max_row, column=2).number_format = number_format
    ws.cell(row=ws.max_row, column=1).font = header_font
    
    # Add net income
    ws.append(["Net Income" if language != 'i' else "Laba Bersih", net_income])
    ws.cell(row=ws.max_row, column=2).number_format = number_format
    ws.cell(row=ws.max_row, column=1).font = header_font
    
    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    
    wb.save(filename)

def export_cash_flow_to_excel(operating_activities, investing_activities, financing_activities, net_cash_flow, filename, language):
    wb = Workbook()
    ws = wb.active
    ws.title = "Cash Flow Statement"
    
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    number_format = '#,##0.00'
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Add headers
    headers = ["Activity", "Amount"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    # Add operating activities
    ws.append(["Operating Activities" if language != 'i' else "Aktivitas Operasi"])
    ws.cell(row=ws.max_row, column=1).font = header_font
    for activity, value in operating_activities.items():
        ws.append([activity, value])
        ws.cell(row=ws.max_row, column=2).number_format = number_format
    
    # Add investing activities
    ws.append(["Investing Activities" if language != 'i' else "Aktivitas Investasi"])
    ws.cell(row=ws.max_row, column=1).font = header_font
    for activity, value in investing_activities.items():
        ws.append([activity, value])
        ws.cell(row=ws.max_row, column=2).number_format = number_format
    
    # Add financing activities
    ws.append(["Financing Activities" if language != 'i' else "Aktivitas Pendanaan"])
    ws.cell(row=ws.max_row, column=1).font = header_font
    for activity, value in financing_activities.items():
        ws.append([activity, value])
        ws.cell(row=ws.max_row, column=2).number_format = number_format
    
    # Add net cash flow
    ws.append(["Net Cash Flow" if language != 'i' else "Arus Kas Bersih", net_cash_flow])
    ws.cell(row=ws.max_row, column=2).number_format = number_format
    ws.cell(row=ws.max_row, column=1).font = header_font
    
    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    
    wb.save(filename)

def export_equity_statement_to_excel(equity_components, total_equity, filename, language):
    wb = Workbook()
    ws = wb.active
    ws.title = "Statement of Changes in Equity"
    
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    number_format = '#,##0.00'
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Add headers
    headers = ["Component", "Amount"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    # Add equity components
    for component, value in equity_components.items():
        ws.append([component, value])
        ws.cell(row=ws.max_row, column=2).number_format = number_format
    
    # Add total equity
    ws.append(["Total Equity" if language != 'i' else "Total Ekuitas", total_equity])
    ws.cell(row=ws.max_row, column=2).number_format = number_format
    ws.cell(row=ws.max_row, column=1).font = header_font
    
    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    
    wb.save(filename)


def export_to_excel(summary, total_sales, total_purchase, net_profit, final_capital, total_assets, total_liabilities, total_equity, filename, language):
    # Convert summary to a DataFrame if it's not already in the correct format
    summary_df = pd.DataFrame.from_dict(summary['sales'], orient='index', columns=['Total_Sales_Amount'])
    summary_df.reset_index(inplace=True)
    summary_df.rename(columns={'index': 'Product_ID'}, inplace=True)

    # Create a workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Financial Report"

    # Define styles
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    number_format = '#,##0.00'
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    if language == 'i':
        # Add column headers
        columns = ["Komponen", "Product ID", "Jumlah"]
        ws.append(columns)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

        # Append rows from DataFrame
        for r in dataframe_to_rows(summary_df, index=False, header=False):
            ws.append([f"Jasa {r[0]}", r[0], r[1]])

        # Format all rows
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
            for cell in row:
                cell.number_format = number_format
                cell.border = border

        # Add Total row
        ws.append(["Total Pendapatan", "", total_sales])
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=2)
        ws["C" + str(ws.max_row)].number_format = number_format
        ws["C" + str(ws.max_row)].border = border

        # Add Financial Summary
        financial_data = [
            ("Laba Rugi Harian", net_profit),
            ("Total Arus Kas", net_profit),
            ("Saldo Kas Harian", final_capital),
            ("Total Aset", total_assets),
            ("Total Utang", total_liabilities),
            ("Total Ekuitas", total_equity)
        ]

        for item in financial_data:
            ws.append(item)
            ws["B" + str(ws.max_row)].number_format = number_format
            ws["B" + str(ws.max_row)].border = border

        # Set column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20

    else:
        # Add column headers
        columns = ["Component", "Product ID", "Amount"]
        ws.append(columns)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

        # Append rows from DataFrame
        for r in dataframe_to_rows(summary_df, index=False, header=False):
            ws.append([f"Service {r[0]}", r[0], r[1]])

        # Format all rows
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
            for cell in row:
                cell.number_format = number_format
                cell.border = border

        # Add Total row
        ws.append(["Total Sales", "", total_sales])
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=2)
        ws["C" + str(ws.max_row)].number_format = number_format
        ws["C" + str(ws.max_row)].border = border

        # Add Financial Summary
        financial_data = [
            ("Daily Profit and Loss", net_profit),
            ("Total Cash Flow", net_profit),
            ("Daily Cash Balance", final_capital),
            ("Total Assets", total_assets),
            ("Total Debt", total_liabilities),
            ("Total Equity", total_equity)
        ]

        for item in financial_data:
            ws.append(item)
            ws["B" + str(ws.max_row)].number_format = number_format
            ws["B" + str(ws.max_row)].border = border

        # Set column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20

    # Save the workbook
    wb.save(filename)

# Balance Sheet, P&L, Cash Flow, Equity, FS, Financial Analysis
# Pilihannya 
